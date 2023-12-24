#!/usr/bin/env python
# coding: utf-8

# In[9]:


import cv2
import re
import numpy as np
import os
import openpyxl
import shutil
from PIL import Image
from pyzbar import pyzbar
from cnocr import CnOcr

# 发票类
class Invoice:
    def __init__(self,payer='',money='',date=''):
        self.payer=payer
        self.money=money
        self.date=date
    # 交易主体、金额、日期
    payer: str
    money:str
    date:str

# 检查是否符合日期格式
def isDate(string):
    date_pattern = r"\d{4}年\d{2}月\d{2}日"
    if re.match(date_pattern, string):
        return True
    else:
        return False
    
# ocr识别接口，参数为图片地址，返回一个Invoice对象        
def ocr_getinfo(image_path):
    # image_path = datab_path + '\\b'+ str(index)+ '.jpg'
    # 旋转图片使得竖图变横图
    image = Image.open(image_path)
    width, height = image.size
    if width < height:
        image_trans = image.transpose(method=Image.ROTATE_90)
        # 如果旋转了，则保存旋转后的图片
        image_trans.save(image_path)
    # 启动cnocr
    ocr = CnOcr()
    out = ocr.ocr(image_path)
    payer = '深圳市购机汇网络有限公司'
    
    invoice = Invoice()
    
    for i,item in enumerate(out):
        text = item['text']
        # 读取付款人
        # 如果当前内容和'深圳市购机汇网络有限公司'至少有八个字符相同，则可认定他们相同
        if len(set(text) & set(payer))>=8:
            invoice.payer = payer
        # 读取日期
        # 如果当前内容是是开票日期，则下一个大概率是日期
        if '开票日期'in text:
            if isDate(out[i+1]['text']):
                invoice.date = out[i+1]['text']
        # 如果当前内容包含2016年，则这个是日期
        if '2016年'in text:
            substring = '2016年'
            index = text.find(substring)
            if index != -1:
                substring_after = text[index + len(substring):]
            invoice.date = substring + substring_after
        # 如果当前内容符合日期格式，则这个是日期
        if isDate(text):
            invoice.date = text
        # 读取金额
        # 如果当前内容包含小写
        if '小写' in text:
            # 如果当前内容包含 . 则说明小写和金额在同一个内容里
            if '.'in text:
                substring = '小写'
                index = text.find(substring)
                if index != -1:
                    substring_after = text[index+4:]
                invoice.money = substring_after
            # 如果当前内容不含 . 则在当前内容的后四个内容里找含 . 的内容
            else:
                for j in range(4):
                    tmp = str(out[i+j]['text'])
                    if '.' in tmp:
                        if tmp[0].isdigit():
                            invoice.money = tmp
                        else:
                            invoice.money = tmp[1:]
        # 对日期的可能错误进行处理
        string = invoice.date
        if(string!=''):
            # 删除空格字符
            string = string.replace(" ", "")
            # 替换特殊字符
            special_chars = {"s":"8","g":"8","E": "日", "e": "8", "l": "1", "I": "1", "i": "1", "o": "0", "O": "0"}
            for char, replacement in special_chars.items():
                string = string.replace(char, replacement)
            invoice.date = string
            
    return invoice


# In[10]:


# 二维码识别结果的返回格式
class Qrcode_return:
    # 是否识别出二维码、发票代码、发票编号、税前金额、日期
    qrcode_exist: bool
    invoice_code: str
    invoice_num:str
    invoice_amount:str
    invoice_date:str
    def __init__(self,qrcode_exist,invoice_code="",invoice_num="",invoice_amount="",invoice_date=""):
        self.qrcode_exist = qrcode_exist
        self.invoice_code = invoice_code
        self.invoice_num = invoice_num
        self.invoice_amount = invoice_amount
        self.invoice_date = invoice_date
        
# 二维码识别接口，参数为图片地址，返回一个Qrcode_return类     
def qrcode_getinfo(image_path):
    image = Image.open(image_path)
    # 竖图旋转成横图
    width, height = image.size
    if width < height:
        image = image.transpose(method=Image.ROTATE_90)
    # 裁剪出二维码部分
    width, height = image.size
    left = 0
    top = height // 10
    right = width // 4
    bottom = height * 2/5
    cropped_image = image.crop((left, top, right, bottom))
    # 将 PIL 图像转换为 OpenCV 图像数组
    image_array = np.array(cropped_image)
    # 将图像数组转换为灰度图像
    gray_image = cv2.cvtColor(image_array, cv2.COLOR_RGB2GRAY)
    # 使用 Otsu's 二值化方法
    _, binary_image = cv2.threshold(gray_image, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    # 使用pyzbar解码二维码
    qrcodes = pyzbar.decode(binary_image,symbols=[pyzbar.ZBarSymbol.QRCODE])
    if len(qrcodes)==0:
        qrcode_return = Qrcode_return(False)
        return qrcode_return;
    else:
        for qrcode in qrcodes:
            string = qrcode.data.decode('utf-8')
        info = string.split(',')
        qrcode_return = Qrcode_return(True,info[2],info[3],info[4],info[5])
        return qrcode_return;


# In[11]:


# 检验该发票金额是否为浮点数，包含对可能错误的处理
# 是浮点数则返回True，不是则返回False
def is_float(invoice):
    string = invoice.money
    # 检查是否符合浮点数的结构，如果可以转换为浮点数，直接返回 True
    try:
        float(string)
        return True
    except ValueError:
        pass
    # 删除空格字符
    string = string.replace(" ", "")
    # 替换特殊字符
    special_chars = {"日": "8", "s":"8","E": "8", "e": "8","g":"8", "l": "1", "I": "1", "i": "1", "o": "0", "O": "0", "口": "0"}
    for char, replacement in special_chars.items():
        string = string.replace(char, replacement)
    invoice.money = string
    # 再试试能不能转换成浮点数
    try:
        float(string)
        print(string)
        return True
    except ValueError:
        return False


# In[12]:


# 校验二维码中的信息与ocr结果,调整识别结果
# 返回True说明无误，返回False说明存在矛盾，需要人工识别
def check_Ocr_Qrcode(invoice,qrcode):
    # 如果二维码中没有信息，返回真
    if(qrcode.invoice_amount==''):
        return True
    # 时间信息以二维码结果为准
    date = qrcode.invoice_date
    formatted_date = date[:4] + "年" + date[4:6] + "月" + date[6:8] + "日"
    invoice.date = formatted_date
    # 金额信息确保二维码的税前金额小于实际金额
    if(qrcode.invoice_amount > invoice.money):
        return False
    else: 
        return True


# In[21]:


# 判断时间是否为2016年6月12日
def check_time(string):
    # 解析年、月、日
    year = int(string[:4])
    month = int(string[5:7])
    day = int(string[8:10])
    # 比较日期
    if (year == 2016 and month == 6 and day == 12):
        return True
    else:
        return False
    
    
# 检查是否通过
def check(invoice):
    # 检查金额是否小于2700
    if float(invoice.money) > 2700:
        return False
    # 检查付款方是不是"深圳市购机汇网络有限公司"
    if invoice.payer!="深圳市购机汇网络有限公司":
        return False
    return check_time(invoice.date)


# # 存入mongodb和数据处理

# In[22]:


import os
from pymongo import MongoClient

# usage:传入保存到的数据库和collection以及要写入的文件位置
# 如imagetomongo = ImageToMongoDB("image", "image_data","./dataset/a")
class ImageToMongoDB:
    def __init__(self, db_name, collection_name):
        self.client = MongoClient('mongodb://localhost:27017/')
        self.db = self.client[db_name]
        self.collection = self.db[collection_name]

    def insert_images(self, folder_path):
        success_count = 0
        failed_files = []

        for root, dirs, files in os.walk(folder_path):
            for file in files:
                if file.lower().endswith('.jpg'):
                    file_path = os.path.join(root, file)
                    try:
                        with open(file_path, 'rb') as f:
                            image_data = f.read()
                            image_doc = {'file_name': file, 'data': image_data}
                            self.collection.insert_one(image_doc)
                            success_count += 1
                    except Exception as e:
                        failed_files.append(file)

        total_processed = success_count + len(failed_files)
        print(f"Total files processed: {total_processed}")
        print(f"Successfully inserted {success_count} files into MongoDB.")
        if failed_files:
            print(f"Failed to insert {len(failed_files)} files into MongoDB: {failed_files}")

# 将每张发票处理结果写入数据库并统计各状态比例以及前k个交易体            
# usage：先实例化eg： data_processor = DataProcessor("ocr_result", "ocr")
# 每次处理后保存状态 eg： status = "通过" data_processor.save_to_mongodb(index, invoice, status)
# 处理完后输出信息以及交易总额前k个交易体eg： top_k = 1  print("前",top_k,"个交易量的交易体为",data_processor.get_top_trading_partners(top_k))  
class DataProcessor:
    def __init__(self, db_name, collection_name):
        self.client = MongoClient('mongodb://localhost:27017/')
        self.db = self.client[db_name]
        self.collection = self.db[collection_name]
        self.total_invoices = 0
        self.approved = 0
        self.rejected = 0
        self.manual_work = 0
        self.trading_partners = {}  # 记录交易方和交易总额的字典

    def save_to_mongodb(self, index, invoice, status):
        money_value = None
        if invoice.money and invoice.money.strip():
            try:
                money_value = float(invoice.money)
            except ValueError:
                pass

        self.total_invoices += 1
        if status == "通过":
            self.approved += 1
            if invoice.payer in self.trading_partners:
                self.trading_partners[invoice.payer] += money_value
            else:
                self.trading_partners[invoice.payer] = money_value
        elif status == "不通过":
            self.rejected += 1
            if invoice.payer in self.trading_partners:
                self.trading_partners[invoice.payer] += money_value
            else:
                self.trading_partners[invoice.payer] = money_value
        elif status == "转人工":
            self.manual_work += 1

        data = {
            "index": index,
            "status": status,
            "details": f"{index} {invoice.payer} {invoice.date} {money_value}"
        }
        self.collection.insert_one(data)

    def get_pass_number(self):
        return self.approved
    
    def get_rejected_number(self):
        return self.rejected
    
    def get_manual_number(self):
        return self.manual_work
    
    def print_info(self):
        print("转人工的个数为：", self.manual_work)  
        print("通过的个数为：", self.approved)  
        print("拒绝的个数为：", self.rejected)  

    def get_top_trading_partners(self, k):
        sorted_partners = sorted(self.trading_partners.items(), key=lambda x: x[1], reverse=True)
        return sorted_partners[:k]


# # 存入excel表格

# In[23]:


import openpyxl

def write_excel(invoice_list: list, index_list: list,save_dir=''):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Sheet1'
    sheet.cell(1, 1, '编号')
    sheet.cell(1, 2, '交易主体')
    sheet.cell(1, 3, '金额')
    sheet.cell(1, 4, '日期')
    for i in range(len(invoice_list)):
        sheet.cell(i+2, 1, index_list[i])
        sheet.cell(i+2, 2, invoice_list[i].payer)
        sheet.cell(i+2, 3, invoice_list[i].money)
        sheet.cell(i+2, 4, invoice_list[i].date)
    # 保存工作簿到指定文件
    wb.save(save_dir)


# # 发邮件

# In[24]:


from smtplib import SMTP_SSL
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.header import Header
import zipfile
import os

token = 'whvohbsoivtkcide'


def zip_dir(dir_path, zip_path):
    zipf = zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED)
    for root, dirs, files in os.walk(dir_path):
        for file in files:
            zipf.write(os.path.join(root, file))
    zipf.close()


def sendEmail(from_addr, to_addr, title, content, img_path = None, excel_path = None, zip_path = None):
    # 1. Create a message
    msg = MIMEMultipart()
    msg['From'] = Header(from_addr)
    msg['To'] = Header(to_addr)
    msg['Subject'] = Header(title)
    msg.attach(MIMEText(content, 'plain', 'utf-8'))

    if img_path:
        with open(img_path, 'rb') as f:
            img_data = f.read()
        att = MIMEApplication(img_data)
        att.add_header('Content-Disposition', 'attachment', filename='invoice.jpg')
        msg.attach(att)

    if excel_path:
        with open(excel_path, 'rb') as f:
            excel_data = f.read()
        att = MIMEApplication(excel_data)
        att.add_header('Content-Disposition', 'attachment', filename='info.xlsx')
        msg.attach(att)

    if zip_path:
        with open(zip_path, 'rb') as f:
            zip_data = f.read()
        att = MIMEApplication(zip_data)
        att.add_header('Content-Disposition', 'attachment', filename='invoice.zip')
        msg.attach(att)

    # 2. Create an SMTP object
    smtp = SMTP_SSL('smtp.qq.com')

    # 3. Login to the server
    smtp.login(from_addr, token)

    # 4. Send email
    smtp.sendmail(from_addr, to_addr, msg.as_string())

    # 5. Close the connection
    smtp.quit()


# # 数据集b处理的完整流程

# In[25]:


# 流程：使用ocr和qrcode识别所有b数据集的图片
def datab_get_info(datab_folder,email_to_addr):
    
    # 需要进行人工检测的下标数组
    manual_work_index = []
    # 需要放进excel表的invoice列表和index列表
    invoice_list = []
    index_list = []
    # ocr存进mongedb的实例
    data_processor = DataProcessor("ocr_result", "ocr")
    
    for filename in os.listdir(datab_folder):
        if filename.endswith(".jpg")or filename.endswith(".png"):
            # 获取图片地址image_path
            image_path = os.path.join(datab_folder, filename)
            # 获取图片下标index
            file_name_without_extension = os.path.splitext(filename)[0]
            number_part = file_name_without_extension[1:]
            index = int(number_part)
            print(index, end=' ')
            
            # 使用ocr识别图片
            invoice = ocr_getinfo(image_path)
            # 如果invoice结果为空，则将图片翻转180度然后重新识别        
            if(invoice.payer==''and invoice.money=='' and invoice.date==''):
                # 旋转图片
                image = Image.open(image_path)
                image_trans = image.transpose(method=Image.ROTATE_180)
                # 保存旋转后的图片
                image_trans.save(image_path)
            invoice = ocr_getinfo(image_path)
            print(invoice.payer, invoice.date, invoice.money)
            # 检查结果是否需要转人工
            # 1.缺少任意一项
            if(invoice.money==''or invoice.payer==''or invoice.date==''):
                manual_work_index.append(index)
                # 存进mongodb
                status = "转人工"
                data_processor.save_to_mongodb(index, invoice, status)
                print("转人工")
                continue
            # 2.金额不是浮点数
            if(not is_float(invoice)):
                manual_work_index.append(index)
                # 存进mongodb
                status = "转人工"
                data_processor.save_to_mongodb(index, invoice, status)
                print("转人工")
                continue
            
            # 使用qrcode识别图片
            qrcode = qrcode_getinfo(image_path)
            # 如果识别成功
            if(qrcode.invoice_date!=''):
                # 校验二维码中的信息与ocr结果
                check_Ocr_Qrcode(invoice,qrcode)
                
            # 检查是否通过
            if check(invoice):
                # 存进mongodb
                status = "通过"
                data_processor.save_to_mongodb(index, invoice, status)
                invoice_list.append(invoice)
                index_list.append(index)
                print("通过")
            else:
                # 存进mongodb
                status = "不通过"
                data_processor.save_to_mongodb(index, invoice, status)
                invoice_list.append(invoice)
                index_list.append(index)
                print("不通过")
    
    # 把invoice列表存进excel
    write_excel(invoice_list,index_list,'invoices.xlsx')
    
    # 输出统计信息
    data_processor.print_info() 
    top_k = 100  
    pass_number = data_processor.get_pass_number()
    rejected_number = data_processor.get_rejected_number()
    manual_number = data_processor.get_manual_number()
    print("前",top_k,"个交易量的交易体为",data_processor.get_top_trading_partners(top_k))
    
    # 把excel发送邮箱
    from_addr = 'snitchyang@qq.com'
    to_addr = email_to_addr
    title = '数据集b结果'
    content = '通过数：' + str(pass_number) + ",不通过数："+ str(rejected_number) + ",转人工数：" + str(manual_number)
    excel_path = 'invoices.xlsx'
    sendEmail(from_addr, to_addr, title,content,None,excel_path,None)
    
    # 把所有需要转人工的图片打包压缩发送邮箱
    # 把所有需要转人工的图片存在新文件夹里
    dir_path = 'manual'
    zip_path = 'manual2.zip'
    # 如果'manual'文件夹不存在，则创建它
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)
    for index in  manual_work_index:
        # 原始图片文件路径
        img_dir = datab_folder + '\\b'+str(index)+'.jpg'
        # 目标图片文件夹
        target_folder = dir_path
        # 将文件从源文件夹复制到目标文件夹
        shutil.copy2(img_dir,target_folder)
    # 压缩文件夹        
    zip_dir(dir_path,zip_path)
    # 发送到邮箱
    from_addr = 'snitchyang@qq.com'
    to_addr = email_to_addr
    title = '需人工处理的图片'
    content = '需人工处理的图片共有'+str(len(manual_work_index))+'个'
    sendEmail(from_addr, to_addr, title,content,None,None,zip_path)
    
    # 把所有图片存到数据库里
    image_manager = ImageToMongoDB('image', 'image_data')
    folder_path = datab_folder
    image_manager.insert_images(folder_path)


# In[20]:


datab_folder = 'C:\\Users\\Max\\Desktop\\发票-大数据集\\发票-大数据集\\b'
email_to_addr = '1276792195@qq.com'
datab_get_info(datab_folder,email_to_addr)

