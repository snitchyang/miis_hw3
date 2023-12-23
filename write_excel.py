import openpyxl


class Invoice:
    def __init__(self,payer='',money='',date=''):
        self.payer=payer
        self.money=money
        self.date=date
    # 交易主体、金额、日期
    payer: str
    money:str
    date:str


def write_excel(invoice_list: list):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Sheet1'
    sheet.cell(1, 1, '交易主体')
    sheet.cell(1, 2, '金额')
    sheet.cell(1, 3, '日期')
    for i in range(len(invoice_list)):
        sheet.cell(i+2, 1, invoice_list[i].payer)
        sheet.cell(i+2, 2, invoice_list[i].money)
        sheet.cell(i+2, 3, invoice_list[i].date)


if __name__ == '__main__':
    write_excel()